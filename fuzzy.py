import openpyxl

# --- Membaca Data dari restoran.xlsx ---
def baca_data(filename):
    """
    Membaca data restoran dari file Excel dan menyimpannya dalam list.
    
    Parameter:
    filename (str): Nama file Excel yang berisi data restoran.
    
    Returns:
    list: Daftar data restoran dengan atribut ID, Kualitas Servis, dan Harga.
    """
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header
        id_restoran, kualitas_servis, harga = row
        data.append({
            'id': id_restoran,
            'kualitas_servis': kualitas_servis,
            'harga': harga
        })
    return data

# --- Fuzzification ---
def fuzzifikasi_servis(kualitas):
    """
    Mengubah kualitas servis restoran (1-100) menjadi nilai fuzzy.
    
    Parameter:
    kualitas (int): Nilai kualitas servis restoran.
    
    Returns:
    dict: Nilai keanggotaan fuzzy untuk kategori Buruk, Sedang, dan Bagus.
    """
    if kualitas <= 40:
        return {'Buruk': 1, 'Sedang': 0, 'Bagus': 0}
    elif 40 < kualitas <= 70:
        return {
            'Buruk': (70 - kualitas) / 30,
            'Sedang': (kualitas - 40) / 30,
            'Bagus': 0
        }
    elif 70 < kualitas <= 90:
        return {
            'Buruk': 0,
            'Sedang': (90 - kualitas) / 20,
            'Bagus': (kualitas - 70) / 20
        }
    else:  # kualitas > 90
        return {'Buruk': 0, 'Sedang': 0, 'Bagus': 1}

def fuzzifikasi_harga(harga):
    """
    Mengubah harga restoran (25.000-55.000) menjadi nilai fuzzy untuk kategori Murah, Sedang, dan Mahal.
    
    Parameter:
    harga (float): Nilai harga restoran.
    
    Returns:
    dict: Nilai keanggotaan fuzzy untuk kategori Murah, Sedang, dan Mahal.
    """
    if harga <= 30000:
        return {'Murah': 1.0, 'Sedang': 0.0, 'Mahal': 0.0}
    elif 30000 < harga <= 40000:
        return {
            'Murah': (40000 - harga) / 10000,
            'Sedang': (harga - 30000) / 10000,
            'Mahal': 0.0
        }
    elif 40000 < harga <= 50000:
        return {
            'Murah': 0.0,
            'Sedang': (50000 - harga) / 10000,
            'Mahal': (harga - 40000) / 10000
        }
    else:  # harga > 50000
        return {'Murah': 0.0, 'Sedang': 0.0, 'Mahal': 1.0}

# --- Inferensi ---
def inferensi(servis_fuzzy, harga_fuzzy):
    """
    Menentukan nilai kelayakan berdasarkan aturan inferensi yang melibatkan kualitas servis dan harga restoran.
    
    Parameter:
    servis_fuzzy (dict): Hasil fuzzifikasi kualitas servis restoran.
    harga_fuzzy (dict): Hasil fuzzifikasi harga restoran.
    
    Returns:
    list: Nilai kelayakan untuk setiap aturan yang diterapkan (9 aturan).
    """
    aturan = [
        # Bagus × [Murah, Sedang, Mahal]
        min(servis_fuzzy['Bagus'], harga_fuzzy['Murah']),   # Aturan 1: Layak Tinggi
        min(servis_fuzzy['Bagus'], harga_fuzzy['Sedang']),  # Aturan 2: Layak Sedang
        min(servis_fuzzy['Bagus'], harga_fuzzy['Mahal']),   # Aturan 3: Layak Rendah
        
        # Sedang × [Murah, Sedang, Mahal]
        min(servis_fuzzy['Sedang'], harga_fuzzy['Murah']),  # Aturan 4: Layak Sedang
        min(servis_fuzzy['Sedang'], harga_fuzzy['Sedang']), # Aturan 5: Layak Sedang
        min(servis_fuzzy['Sedang'], harga_fuzzy['Mahal']),  # Aturan 6: Layak Rendah
        
        # Buruk × [Murah, Sedang, Mahal]
        min(servis_fuzzy['Buruk'], harga_fuzzy['Murah']),   # Aturan 7: Layak Rendah
        min(servis_fuzzy['Buruk'], harga_fuzzy['Sedang']),  # Aturan 8: Layak Sangat Rendah
        min(servis_fuzzy['Buruk'], harga_fuzzy['Mahal'])    # Aturan 9: Layak Sangat Rendah
    ]
    return aturan

# --- Defuzzification ---
def defuzzifikasi(aturan):
    """
    Mengubah hasil inferensi fuzzy menjadi skor numerik menggunakan metode Weighted Average (rata-rata terbobot).
    
    Parameter:
    aturan (list): Nilai kekuatan untuk setiap aturan yang diterapkan (9 aturan).
    
    Returns:
    float: Skor kelayakan restoran (0-100).
    """
    # Bobot untuk setiap aturan (sesuai dengan urutan aturan di atas)
    bobot_kelayakan = [
        100,  # Aturan 1: Bagus & Murah → Layak Tinggi (diubah dari 90 ke 100)
        75,   # Aturan 2: Bagus & Sedang → Layak Sedang
        50,   # Aturan 3: Bagus & Mahal → Layak Rendah
        75,   # Aturan 4: Sedang & Murah → Layak Sedang
        60,   # Aturan 5: Sedang & Sedang → Layak Sedang
        40,   # Aturan 6: Sedang & Mahal → Layak Rendah
        50,   # Aturan 7: Buruk & Murah → Layak Rendah
        30,   # Aturan 8: Buruk & Sedang → Layak Sangat Rendah
        20    # Aturan 9: Buruk & Mahal → Layak Sangat Rendah
    ]
    # Hitung rata-rata terbobot
    pembilang = sum(a * b for a, b in zip(aturan, bobot_kelayakan))
    penyebut = sum(aturan)

    # Hindari pembagian dengan nol
    if penyebut == 0:
        return 0.0
    return pembilang / penyebut

# --- Menyimpan Output ke Peringkat.xlsx ---
def simpan_output(filename, hasil):
    """
    Menyimpan hasil peringkat restoran ke dalam file Excel.
    
    Parameter:
    filename (str): Nama file tempat hasil disimpan.
    hasil (list): Daftar hasil peringkat restoran dengan skor kelayakan.
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['ID', 'Kualitas Servis', 'Harga', 'Skor Kelayakan'])

    for item in hasil:
        sheet.append([item['id'], item['kualitas_servis'], item['harga'], item['skor']])

    wb.save(filename)

# --- Program Utama ---
if __name__ == "__main__":
    # Membaca data restoran dari file restoran.xlsx
    data = baca_data('restoran.xlsx')

    hasil = []
    for d in data:
        servis_fuzzy = fuzzifikasi_servis(d['kualitas_servis'])
        harga_fuzzy = fuzzifikasi_harga(d['harga'])
        aturan = inferensi(servis_fuzzy, harga_fuzzy)
        skor = defuzzifikasi(aturan)

        hasil.append({
            'id': d['id'],
            'kualitas_servis': d['kualitas_servis'],
            'harga': d['harga'],
            'skor': skor,
            'aturan': aturan,
            'servis_fuzzy': servis_fuzzy,  # Simpan hasil fuzzifikasi
            'harga_fuzzy': harga_fuzzy     # Simpan hasil fuzzifikasi
        })

    # Sortir hasil berdasarkan skor kelayakan tertinggi
    hasil.sort(key=lambda x: x['skor'], reverse=True)

    # Ambil 3 restoran terbaik (top 3) untuk tracing aturan inferensi
    top_5 = hasil[:5]

    # Simpan hasil ke dalam file peringkat.xlsx
    simpan_output('peringkat.xlsx', hasil[:10])  # Simpan 10 restoran terbaik

    # Tampilkan hasil dalam format tabel dengan tracing untuk top 3
    print(f"{'No.':<4} {'ID':<8} {'Kualitas Servis':<18} {'Harga':<15} {'Skor Kelayakan':<20}")
    print("="*70)
    
    for i, item in enumerate(top_5, 1):
        print(f"\nDetail Proses Fuzzy Logic untuk Restoran ID {item['id']}:")
        print("="*70)
        print(f"ID Restoran: {item['id']}")
        print(f"Kualitas Servis: {item['kualitas_servis']} -> {item['servis_fuzzy']}")
        print(f"Harga: Rp {item['harga']} -> {item['harga_fuzzy']}")
        print("\nHasil Inferensi (9 aturan):")  # Update ke 9 aturan
        for idx, val in enumerate(item['aturan'], 1):
            print(f"Aturan {idx}: {val:.2f}")
        
        # Skor akhir dari defuzzifikasi
        print(f"\nSkor Akhir (Defuzzifikasi): {item['skor']:.2f}")
        print("="*70)

    # Tampilkan tabel 10 restoran terbaik
    print("\nDaftar 10 Restoran Terbaik berdasarkan Fuzzy Logic:")
    print(f"{'No.':<4} {'ID':<8} {'Kualitas Servis':<18} {'Harga':<15} {'Skor Kelayakan':<20}")
    print("="*70)
    
    for i, item in enumerate(hasil[:10], 1):
        print(f"{i:<4} {item['id']:<8} {item['kualitas_servis']:<18} {item['harga']:<15} {item['skor']:<20.2f}")
